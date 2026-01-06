from playwright.sync_api import sync_playwright
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()


def main():
    def next_tab(number):
        for _ in range(number):
            page.get_by_role("button", name="Próximo").click()
            page.wait_for_timeout(TIMEOUT)

    with sync_playwright() as p:
        TIMEOUT = 2000
        EXCEL_NAME = "estagiarios.xlsx"
        LASTNAME = "Estagiário"
        PASSWORD = "Primeiro@01"

        browser = p.chromium.launch(
            headless=False, slow_mo=50, args=['--start-maximized'])
        page = browser.new_page()
        page.goto(os.getenv("ZIMBRA_ADMIN_URL"))

        page.get_by_role("textbox", name="Username: Username:").fill(
            os.getenv("ZIMBRA_ADMIN_USER"))
        page.get_by_role("textbox", name="Password: Password:").fill(
            os.getenv("ZIMBRA_ADMIN_PASS"))
        page.get_by_role("button", name="Sign In").click()
        page.wait_for_timeout(TIMEOUT)

        excel_dir = os.path.join(os.getcwd(), EXCEL_NAME)
        wb = load_workbook(excel_dir, data_only=True)
        excel = wb.active

        for row in excel.iter_rows(min_row=2, max_col=3, max_row=excel.max_row):
            nome, lotacao, funcional = row

            funcional_formatada = f"e.{funcional.value}"

            page.locator("[id=\"_XForm_query_display\"]").fill(
                funcional_formatada)
            page.keyboard.press("Enter")
            page.wait_for_timeout(TIMEOUT)
            if page.get_by_text("Não foi encontrado nenhum resultado.").count() > 0:
                page.goto(os.getenv("ZIMBRA_ADMIN_URL"))
                page.wait_for_timeout(TIMEOUT)
                page.get_by_text("Adicionar conta...").click()
                page.wait_for_timeout(TIMEOUT)
                page.locator("#zdlgv__NEW_ACCT_name_2").fill(
                    funcional_formatada)
                page.wait_for_timeout(TIMEOUT)
                page.locator("#zdlgv__NEW_ACCT_givenName").fill(
                    nome.value.title())
                page.wait_for_timeout(TIMEOUT)
                page.locator("#zdlgv__NEW_ACCT_sn").fill(LASTNAME)
                page.wait_for_timeout(TIMEOUT)
                page.locator("#zdlgv__NEW_ACCT_password").fill(PASSWORD)
                page.wait_for_timeout(TIMEOUT)
                page.locator("#zdlgv__NEW_ACCT_confirmPassword").fill(PASSWORD)
                page.wait_for_timeout(TIMEOUT)
                page.locator(
                    "#zdlgv__NEW_ACCT_zimbraPasswordMustChange").check()
                page.locator("#zdlgv__NEW_ACCT_description").fill(
                    f"{LASTNAME} {lotacao.value.title()}")
                page.wait_for_timeout(TIMEOUT)

                # Aba Membro de
                next_tab(2)
                page.get_by_role("button", name="Buscar").click()
                page.wait_for_timeout(TIMEOUT)
                page.get_by_role("cell", name="newsletter_@pci.es.gov.br",
                                 exact=True).click()
                page.wait_for_timeout(TIMEOUT)
                page.get_by_role("button", name="Adicionar",
                                 exact=True).click()
                page.wait_for_timeout(TIMEOUT)

                # Aba Avançado
                next_tab(6)
                page.locator(
                    "#zdlgv__NEW_ACCT_zimbraPasswordMinLength_2").fill("8")
                page.wait_for_timeout(TIMEOUT)
                page.locator(
                    "#zdlgv__NEW_ACCT_zimbraPasswordMaxLength_2").fill("30")
                page.wait_for_timeout(TIMEOUT)
                page.locator(
                    "#zdlgv__NEW_ACCT_zimbraPasswordMinUpperCaseChars_2").fill("1")
                page.wait_for_timeout(TIMEOUT)
                page.locator("#zdlgv__NEW_ACCT_zimbraPasswordMinPunctuationChars_2").fill(
                    "1")
                page.wait_for_timeout(TIMEOUT)
                page.locator(
                    "#zdlgv__NEW_ACCT_zimbraPasswordMinNumericChars_2").fill("1")
                page.wait_for_timeout(TIMEOUT)
                page.locator(
                    "#zdlgv__NEW_ACCT_zimbraPasswordMaxAge_2").fill("90")
                page.wait_for_timeout(TIMEOUT)
                page.get_by_role("button", name="Concluir").click()
                page.wait_for_timeout(TIMEOUT)
                print(f"Conta {funcional_formatada} cadastrada com sucesso")
            else:
                print(f"Conta {funcional_formatada} já existe")
                continue

        page.pause()


if __name__ == "__main__":
    main()
